#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Compile a Bao Gia summary index from folder/file names."""
import os, re, sys, unicodedata
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

# ---------------- CONFIG ----------------
READABLE_EXT = {'.pdf', '.xls', '.xlsx', '.xlsm', '.doc', '.docx',
                '.jpg', '.jpeg', '.png', '.ppt', '.pptx'}
DELIM = '@'

# order number -> (formal name, is_special). Roman derived from order.
CATEGORIES = {
    1:  ('Công bố giá',     True),
    2:  ('Cước vận chuyển', True),
    3:  ('Kết cấu',         False),
    4:  ('Kiến trúc',       False),
    5:  ('Chống mối',       False),
    6:  ('Điện',            False),
    7:  ('Điện nhẹ',        False),
    8:  ('CTN',             False),
    9:  ('DHTG',            False),
    10: ('PCCC',            False),
    11: ('BG Thang may',    False),
    12: ('Hạ Tầng',         False),
}

CHOSEN_KEYS    = {'chon','c','co','x','v','ok','yes'}
NOTCHOSEN_KEYS = {'khongchon','kochon','ko','khong','k','kdung','khongdung','kodung','no','n'}

NORMALIZE_PARTNER_PREFIX = True  # "Cong ty"/"cty" -> "Công ty"
# leading "quotation" words stripped from the subcategory (diacritic-insensitive)
QUOTE_PREFIXES = ['bao gia', 'bang gia', 'bg']
CAPITALIZE_SUBCAT = True
# Tieu de & thong tin du an (co the dat trong cau_hinh.txt, muc [THONG TIN])
TITLE_MAIN = 'DANH MỤC BÁO GIÁ'
PROJECT_NAME = ''   # DỰ ÁN
WORK_NAME = ''      # CÔNG TRÌNH
LOCATION = ''       # ĐỊA ĐIỂM
HEADER_FILL = 'FFB4C7E7'   # xanh nhat (Blue, Accent 1, Lighter 60%)
# ----------------------------------------

def strip_diacritics(s):
    return ''.join(c for c in unicodedata.normalize('NFD', s)
                   if unicodedata.category(c) != 'Mn').replace('đ','d').replace('Đ','D')

def status_key(s):
    return re.sub(r'[^a-z0-9]', '', strip_diacritics(s).lower())

def int_to_roman(n):
    vals=[(1000,'M'),(900,'CM'),(500,'D'),(400,'CD'),(100,'C'),(90,'XC'),
          (50,'L'),(40,'XL'),(10,'X'),(9,'IX'),(5,'V'),(4,'IV'),(1,'I')]
    out=''
    for v,sym in vals:
        while n>=v: out+=sym; n-=v
    return out

def clean_field(s):
    s = s.strip()
    # strip one layer of wrapping brackets/quotes
    pairs = [('[',']'),('(',')'),('{','}'),('"','"'),("'","'")]
    changed = True
    while changed:
        changed = False
        for a,b in pairs:
            if len(s) >= 2 and s.startswith(a) and s.endswith(b):
                s = s[1:-1].strip(); changed = True
    # strip leading/trailing stray punctuation/dashes
    s = s.strip(' -–—_,.;:')
    s = re.sub(r'\s+', ' ', s)
    return s.strip()

def strip_quote_prefix(s):
    nk = strip_diacritics(s).lower()
    for pref in QUOTE_PREFIXES:
        if nk.startswith(pref) and (len(nk) == len(pref) or nk[len(pref)] in ' -–—_:.'):
            rem = s[len(pref):].lstrip(' -–—_:.')
            if rem.strip() and not re.fullmatch(r'[\W\d]+', rem.strip()):
                return rem.strip()
    return s

def cap_first(s):
    if not s or not CAPITALIZE_SUBCAT:
        return s
    for i, ch in enumerate(s):
        if ch.isalpha():
            return s[:i] + ch.upper() + s[i+1:]
    return s

def normalize_partner(p):
    if not p or not NORMALIZE_PARTNER_PREFIX:
        return p or ''
    toks = p.split()
    if not toks:
        return p
    low = [strip_diacritics(t).lower().strip('.:-') for t in toks]
    if low[0] in ('cty', 'c.ty', 'congty'):
        return ('Công ty ' + ' '.join(toks[1:])).strip()
    if len(low) >= 2 and low[0] == 'cong' and low[1] == 'ty':
        return ('Công ty ' + ' '.join(toks[2:])).strip()
    return p

def parse_status(raw):
    """returns (display, warning_or_None)"""
    k = status_key(raw)
    if not k: return (None, None)
    if k in CHOSEN_KEYS: return ('Chọn', None)
    if k in NOTCHOSEN_KEYS: return ('Không chọn', None)
    return (None, f'không nhận diện trạng thái: "{raw}"')

def parse_name(stem):
    """stem = filename without extension. returns dict or None."""
    if DELIM not in stem:
        return None
    segs = [clean_field(x) for x in stem.split(DELIM)]
    segs = [s for s in segs if s != '']
    if not segs:
        return None
    subcat = cap_first(strip_quote_prefix(segs[0]))
    rest = segs[1:]
    partner, status, warn = '', None, None
    if rest:
        st, w = parse_status(rest[-1])
        if st is not None:
            status = st
            partner = ' '.join(rest[:-1])
        elif len(rest) >= 2:
            # positional status slot but unrecognized -> keep raw, warn
            status = None; warn = w
            partner = ' '.join(rest[:-1])
            # still surface the raw token by appending? keep blank, warn only
            warn = f'không nhận diện trạng thái: "{rest[-1]}"'
        else:
            partner = ' '.join(rest)
    partner = normalize_partner(partner.strip())
    return {'subcat': subcat, 'partner': partner, 'status': status, 'warn': warn}

def scan(project_dir):
    """returns categories list + skipped + warnings"""
    entries = sorted(os.listdir(project_dir))
    # Chi nhan thu muc co dang "<so><dau phan cach>..." (vd "3. ...", "5 - ...").
    # Moi thu muc khac (tham khao, ten bat dau bang @ ! chu cai...) deu BO QUA.
    folders_by_order = {}   # so -> list of (folder_name, desc)
    for name in entries:
        full = os.path.join(project_dir, name)
        if not os.path.isdir(full):
            continue
        # Chap nhan "4. Ten", "15. Ten"... nhung TU CHOI "4.6.2026 ..." (ngay thang):
        # sau dau phan cach phai la ky tu KHONG phai chu so (hoac het ten).
        m = re.match(r'\s*(\d+)\s*[.\-_)]\s*(?=\D|$)(.*)', name)
        if not m:
            continue  # khong co so o dau -> bo qua hoan toan
        folders_by_order.setdefault(int(m.group(1)), []).append((name, m.group(2).strip()))

    skipped, warnings = [], []
    cats = []  # (order, roman, formal, special, desc_name, items)

    def read_records(folder_name, special):
        records = []
        full = os.path.join(project_dir, folder_name)
        if not os.path.isdir(full):
            return records
        for fn in sorted(os.listdir(full)):
            fpath = os.path.join(full, fn)
            if not os.path.isfile(fpath):
                continue
            stem, ext = os.path.splitext(fn)
            if ext.lower() not in READABLE_EXT:
                continue  # bo qua hoan toan (zip, .hce, Thumbs.db, ...)
            parsed = parse_name(stem)
            if parsed is None:
                reason = ('tệp gốc trong nhóm I/II (xem ghi chú)' if special
                          else 'không theo quy ước đặt tên')
                skipped.append((folder_name, fn, reason))
                continue
            if parsed['warn']:
                warnings.append((folder_name, fn, parsed['warn']))
            records.append(parsed)
        return records

    def build_items(records, folder_label):
        groups, order_seen, normmap = {}, [], {}
        for r in records:
            key = r['subcat']
            if key not in groups:
                groups[key] = []; order_seen.append(key)
                normmap.setdefault(status_key(key), []).append(key)
            groups[key].append(r)
        for nk, variants in normmap.items():
            if len(variants) > 1:
                warnings.append((folder_label, '', 'có thể trùng hạng mục (khác chính tả): '
                                 + ' / '.join(f'"{v}"' for v in variants)))
        rank = {'Chọn': 0, None: 1, 'Không chọn': 2}
        items = []
        for sub in order_seen:
            parts = sorted(groups[sub], key=lambda r: (rank.get(r['status'], 1), (r['partner'] or '')))
            items.append((sub, parts))
        return items

    # --- Bo khung co dinh I..XII (luon hien thi du co thieu thu muc) ---
    for order in sorted(CATEGORIES):
        formal, special = CATEGORIES[order]
        roman = int_to_roman(order)
        lst = folders_by_order.get(order, [])
        desc = (lst[0][1] if (special and lst) else '')
        recs = []
        for (fname, _d) in lst:
            recs += read_records(fname, special)
        label = lst[0][0] if lst else ''
        cats.append((order, roman, formal, special, desc, build_items(recs, label)))

    # --- Thu muc co so NGOAI 1..12 (vd 00, 13, 14...) -> ghep them, ten = ten thu muc ---
    nextn = max(CATEGORIES) + 1
    for o in sorted(x for x in folders_by_order if x not in CATEGORIES):
        for (fname, d) in folders_by_order[o]:
            name = d or fname
            cats.append((nextn, int_to_roman(nextn), name, False, '',
                         build_items(read_records(fname, False), fname)))
            nextn += 1

    # --- Canh bao trung so thu tu thu muc ---
    for o, lst in sorted(folders_by_order.items()):
        if len(lst) > 1:
            warnings.append(('', '', f'TRÙNG số thứ tự thư mục [{o}] ở: '
                             + ', '.join(f'"{n}"' for n, _ in lst)
                             + ' — đã gộp/đưa hết vào báo cáo, vui lòng kiểm tra.'))

    return cats, skipped, warnings

# ---------------- BUILD WORKBOOK ----------------
TNR = 'Times New Roman'
thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def cell(ws, r, c, val, bold=False, align='center', size=12, wrap=False):
    x = ws.cell(r, c, val)
    x.font = Font(name=TNR, size=size, bold=bold)
    x.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    x.border = BORDER
    return x

def build(cats, skipped, warnings, title, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Báo cáo'
    ws.column_dimensions['A'].width = 9.11
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 32
    ws.column_dimensions['D'].width = 14.11

    def banner(row, text, size, height=None):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row, 1, text)
        c.font = Font(name=TNR, size=size, bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')
        if height:
            ws.row_dimensions[row].height = height

    def meta(label, value):
        return f'{label} {value}'.strip() if value else label

    # rows 1-4: tieu de + thong tin du an
    banner(1, TITLE_MAIN, 16, height=24)
    banner(2, meta('DỰ ÁN:', PROJECT_NAME), 12)
    banner(3, meta('CÔNG TRÌNH:', WORK_NAME), 12)
    banner(4, meta('ĐỊA ĐIỂM:', LOCATION), 12)
    # row 5: dong trong (breather)

    # row 6: header row - in hoa, dam, can giua, to mau xanh nhat
    fill = PatternFill('solid', fgColor=HEADER_FILL)
    for j, h in enumerate(['STT', 'NỘI DUNG', 'TÊN CÔNG TY', 'GHI CHÚ'], start=1):
        x = cell(ws, 6, j, h, bold=True, align='center')
        x.fill = fill
    ws.row_dimensions[6].height = 28.2

    # cot A va D can giua; cot B va C can trai
    ALIGN = {1: 'center', 2: 'left', 3: 'left', 4: 'center'}

    r = 7
    for (order, roman, formal, special, desc, items) in cats:
        # dong tieu de nhom lon (in dam)
        cell(ws, r, 1, roman, bold=True, align=ALIGN[1])
        cell(ws, r, 2, formal, bold=True, align=ALIGN[2])
        cell(ws, r, 3, (desc if special else None), bold=True, align=ALIGN[3])
        cell(ws, r, 4, None, bold=True, align=ALIGN[4])
        r += 1
        prev_item_row = None
        for sub, parts in items:
            first_row_of_item = r
            for k, p in enumerate(parts):
                if k == 0:
                    cell(ws, r, 1, 1 if prev_item_row is None else f'=A{prev_item_row}+1', align=ALIGN[1])
                else:
                    cell(ws, r, 1, None, align=ALIGN[1])
                cell(ws, r, 2, sub, align=ALIGN[2])
                cell(ws, r, 3, p['partner'] or None, align=ALIGN[3])
                cell(ws, r, 4, p['status'], align=ALIGN[4])
                r += 1
            prev_item_row = first_row_of_item
    last_row = r - 1

    # --- thiet lap in A4: doc, vua chieu rong 1 trang, can giua ngang, le co dinh ---
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.7, right=0.7, top=0.75, bottom=0.75, header=0.3, footer=0.3)
    ws.print_area = f'A1:D{last_row}'
    ws.print_title_rows = '6:6'  # lap lai dong tieu de cot o moi trang in

    # ---- Data & check sheet ----
    ws2 = wb.create_sheet('Dữ liệu & Kiểm tra')
    for col,w in zip('ABCDEF',[6,18,34,30,12,40]):
        ws2.column_dimensions[col].width = w
    def row2(vals, bold=False, r=None):
        for j,v in enumerate(vals, start=1):
            x = ws2.cell(r, j, v)
            x.font = Font(name=TNR, size=11, bold=bold)
            x.alignment = Alignment(vertical='center', wrap_text=True)
        return r
    rr = 1
    ws2.cell(rr,1,'DỮ LIỆU PHẲNG (mỗi dòng = 1 tệp được đọc)').font = Font(name=TNR,size=11,bold=True)
    rr += 1
    row2(['#','Hạng mục lớn','Nội dung (hạng mục)','Tên công ty','Ghi chú','Tệp nguồn'], bold=True, r=rr); rr+=1
    n=1
    for (order, roman, formal, special, desc, items) in cats:
        for sub,parts in items:
            for p in parts:
                row2([n, f'{roman}. {formal}', sub, p['partner'], p['status'] or '', ''], r=rr)
                rr+=1; n+=1
    rr+=1
    ws2.cell(rr,1,'TỆP BỎ QUA – đúng định dạng đọc nhưng SAI quy ước (cần đặt lại tên nếu cần đọc)').font=Font(name=TNR,size=11,bold=True)
    rr+=1
    row2(['','Thư mục','Tên tệp','Lý do','',''], bold=True, r=rr); rr+=1
    for folder,fn,reason in skipped:
        row2(['', folder, fn, reason,'',''], r=rr); rr+=1
    rr+=1
    ws2.cell(rr,1,'CẢNH BÁO').font=Font(name=TNR,size=11,bold=True); rr+=1
    if not warnings:
        row2(['','(không có)','','','',''], r=rr); rr+=1
    for folder,fn,w in warnings:
        row2(['', folder, fn, w,'',''], r=rr); rr+=1

    wb.save(out_path)

def _norm_header(s):
    return re.sub(r'[^a-z]', '', strip_diacritics(s).lower())

def _truthy(v):
    return strip_diacritics(v).strip().lower() in {'co','yes','true','x','1','bat','on','y'}

def load_config(path):
    """Override defaults from a plain-text config if present. Tolerant of mistakes."""
    global CATEGORIES, CHOSEN_KEYS, NOTCHOSEN_KEYS, QUOTE_PREFIXES, READABLE_EXT
    global NORMALIZE_PARTNER_PREFIX, CAPITALIZE_SUBCAT
    global TITLE_MAIN, PROJECT_NAME, WORK_NAME, LOCATION
    if not path or not os.path.isfile(path):
        return False
    cats, chosen, notchosen, prefixes, exts = {}, [], [], [], []
    section = None
    try:
        with open(path, encoding='utf-8-sig') as f:
            for raw in f:
                line = raw.strip()
                if not line or line.startswith('#'):
                    continue
                if line.startswith('[') and line.endswith(']'):
                    section = _norm_header(line); continue
                if section == 'danhmuc':
                    parts = [p.strip() for p in line.split('|')]
                    if len(parts) >= 2 and parts[0].isdigit():
                        special = _truthy(parts[2]) if len(parts) >= 3 else False
                        cats[int(parts[0])] = (parts[1], special)
                elif section == 'chon':
                    chosen += [x.strip() for x in line.split(',') if x.strip()]
                elif section == 'khongchon':
                    notchosen += [x.strip() for x in line.split(',') if x.strip()]
                elif section == 'botiento':
                    prefixes += [x.strip() for x in line.split(',') if x.strip()]
                elif section in ('duoifile', 'loaifile'):
                    exts += [x.strip() for x in line.split(',') if x.strip()]
                elif section == 'tuychon' and '=' in line:
                    k, v = line.split('=', 1); kk = _norm_header(k)
                    if 'congty' in kk:
                        NORMALIZE_PARTNER_PREFIX = _truthy(v)
                    elif 'viethoa' in kk or 'hoachu' in kk:
                        CAPITALIZE_SUBCAT = _truthy(v)
                elif section == 'thongtin' and '=' in line:
                    k, v = line.split('=', 1); kk = _norm_header(k); v = v.strip()
                    if 'tieude' in kk and v:
                        TITLE_MAIN = v
                    elif 'duan' in kk:
                        PROJECT_NAME = v
                    elif 'congtrinh' in kk:
                        WORK_NAME = v
                    elif 'diadiem' in kk:
                        LOCATION = v
    except Exception as e:
        print('Khong doc duoc cau_hinh.txt, dung mac dinh. Chi tiet:', e)
        return False
    if cats: CATEGORIES = cats
    if chosen: CHOSEN_KEYS = set(filter(None, (status_key(x) for x in chosen)))
    if notchosen: NOTCHOSEN_KEYS = set(filter(None, (status_key(x) for x in notchosen)))
    if prefixes: QUOTE_PREFIXES = [strip_diacritics(x).lower().strip() for x in prefixes]
    if exts: READABLE_EXT = set('.' + e.lstrip('.').lower() for e in exts)
    return True

if __name__ == '__main__':
    project = os.path.normpath(sys.argv[1].strip().strip('"'))
    out = (sys.argv[2].strip().strip('"') if len(sys.argv) > 2 else 'output.xlsx')
    if not os.path.isdir(project):
        print('Khong tim thay thu muc:', project)
        sys.exit(1)
    cfg = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'cau_hinh.txt')
    if load_config(cfg):
        print('Da doc cau hinh tu: cau_hinh.txt')
    else:
        print('Dung cau hinh mac dinh (khong thay cau_hinh.txt).')
    base = os.path.basename(project)
    title = 'BẢNG TỔNG HỢP BÁO GIÁ – ' + re.sub(r'^\s*\d+\s*[.\-_)]\s*','',base)
    cats, skipped, warnings = scan(project)
    build(cats, skipped, warnings, title, out)
    print('Categories:', len(cats),
          '| items:', sum(len(i[5]) for i in cats),
          '| skipped:', len(skipped), '| warnings:', len(warnings))
